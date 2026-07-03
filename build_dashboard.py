#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
经营驾驶舱 一键更新脚本
使用方法：
  1. 将本脚本与5张Excel文件、dashboard_template.html 放在同一目录下
  2. 安装依赖：pip install openpyxl
  3. 运行：python build_dashboard.py
  4. 生成：index.html（GitHub Pages 首页，直接用浏览器打开即可）

每次更新Excel后重新运行本脚本即可。
"""

import os
import sys
import json
import shutil
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("[错误] 缺少依赖：openpyxl，请运行：pip install openpyxl")
    sys.exit(1)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
def path(name): return os.path.join(BASE_DIR, name)

PRODUCT_IMAGE_DIR = Path(r"D:\总经办工作文档\商品图片\（提取图片）产品图册")
PUBLIC_PRODUCT_IMAGE_DIR = Path(BASE_DIR) / "assets" / "product-images"

EXCEL_FILES = ["销售数据.xlsx","库存数据.xlsx","指标.xlsx","商品物料.xlsx","仓库与店铺对应关系表.xlsx"]
for f in EXCEL_FILES:
    if not os.path.exists(path(f)):
        print(f"[错误] 找不到文件：{f}")
        sys.exit(1)

TEMPLATE = "dashboard_template.html"
if not os.path.exists(path(TEMPLATE)):
    print(f"[错误] 找不到模板文件：{TEMPLATE}（需与本脚本放在同一目录）")
    sys.exit(1)

print("=" * 60)
print("经营驾驶舱 数据更新中，请稍候...")
print("=" * 60)

# =========================================================
# 1. 店铺 → 客户 映射
# =========================================================
STORE_MAPPING = {
    "DAPHNE上海奉贤金汇天街直营店":"DA奉贤金汇",
    "DAPHNE上海普陀中环百联直营店":"DA中环百联",
    "【得物】DAPHNE.LAB商城":"得物",
    "【得物】DAPHNE.LAB商城(新)":"得物",
    "【抖音】DAPHNE LAB女鞋旗舰店":"抖音",
    "【抖音】DAPHNE.LAB官方旗舰店":"抖音",
    "【京东】DAPHNE.LAB京东自营旗舰店":"京东",
    "【京东】DAPHNE.LAB旗舰店":"京东",
    "【天猫】DAPHNE.LAB旗舰店":"天猫",
    "【唯品会】DAPHNE.LAB官方特卖旗舰店":"唯品会",
    "【小红书】DAPHNE.LAB旗舰店":"小红书",
    "tube-白噪音":"其他","tube-圣古屋":"其他",
    "北京-艾立朋":"其他","河南-张建乐":"其他","陕西-窦福轩":"其他",
    "上海波妞网络科技有限公司":"其他","线下渠道":"其他",
    "DAPHNE.LAB成都COSMO直营店":"成都COSMO",
    "DAPHNE.LAB南京艾尚天地直营店":"南京IST",
    "DAPHNE.LAB上海TX淮海直营店":"TX淮海",
    "DAPHNE.LAB上海淮海百盛直营店":"淮海百盛",
    "DAPHNE.LAB长沙TX直营店":"TX长沙",
    "DAPHNE.LAB TX长沙直营店":"TX长沙",
}

# 以“仓库与店铺对应关系表.xlsx”D/E列为最新店铺映射，覆盖脚本默认值。
mapping_wb = openpyxl.load_workbook(path("仓库与店铺对应关系表.xlsx"), read_only=True, data_only=True)
mapping_ws = mapping_wb["Sheet1"]
mapping_store_count = 0
for r in mapping_ws.iter_rows(min_row=2, values_only=True):
    store = r[3] if len(r) > 3 else None
    customer = r[4] if len(r) > 4 else None
    if store and customer:
        STORE_MAPPING[str(store).strip()] = str(customer).strip()
        mapping_store_count += 1
mapping_wb.close()
print(f"店铺映射表已加载：{mapping_store_count} 条")

CHANNEL_GROUPS = {
    "DA": {"DA奉贤金汇","DA中环百联"},
    "DL线上": {"得物","抖音","京东","天猫","唯品会","小红书"},
    "DL线下": {"成都COSMO","南京IST","TX淮海","淮海百盛","TX长沙"},
    "其他": {"其他"},
}
def chan_of(cust):
    for g, s in CHANNEL_GROUPS.items():
        if cust in s: return g
    return "其他"

# =========================================================
# 2. 商品物料：SKC 元数据
# =========================================================
print("[1/6] 读取商品物料...")
TARGET_CATS_5 = {"女鞋","服装","箱包","配件","赠品"}
TARGET_CATS_3 = {"女鞋","服装","箱包"}
SEASON_ORDER = {"春":0,"夏":1,"秋":2,"冬":3}

skc_meta = {}  # skc -> {cat, small, nianji, factory, name}
wb = openpyxl.load_workbook(path("商品物料.xlsx"), read_only=True, data_only=True)
for r in wb["Sheet1"].iter_rows(min_row=2, values_only=True):
    style,skc,big,mid,small,name,brand,nianji,year,season,price,factory = r
    if big in TARGET_CATS_5:
        skc_meta[skc] = {
            "cat": big, "small": small or "未知小类",
            "nianji": str(nianji) if nianji else "未知年季",
            "factory": (factory or "未知工厂").strip(),
            "name": name or "",
        }

skc_cat  = {k: v["cat"]     for k, v in skc_meta.items()}
skc_name = {k: v["name"]    for k, v in skc_meta.items() if v["cat"] in TARGET_CATS_3}
skc_factory = {k: v["factory"] for k, v in skc_meta.items()}
skc_nianji = {k: v["nianji"] for k, v in skc_meta.items()}
print(f"    商品物料 SKC 数：{len(skc_meta)}")

# 销售排行商品图：复制到看板目录，HTML 使用网页可访问的相对 URL。
skc_image = {}
if PRODUCT_IMAGE_DIR.is_dir():
    PUBLIC_PRODUCT_IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    copied_count = 0
    for image_file in PRODUCT_IMAGE_DIR.iterdir():
        if image_file.is_file() and image_file.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp"}:
            skc = image_file.stem
            if skc in skc_meta and skc not in skc_image:
                public_file = PUBLIC_PRODUCT_IMAGE_DIR / image_file.name
                if (not public_file.exists()
                        or public_file.stat().st_size != image_file.stat().st_size
                        or public_file.stat().st_mtime < image_file.stat().st_mtime):
                    shutil.copy2(image_file, public_file)
                    copied_count += 1
                skc_image[skc] = f"assets/product-images/{image_file.name}"
    print(f"    商品图片匹配：{len(skc_image)}/{len(skc_meta)}，本次复制/更新：{copied_count}")
else:
    print(f"[提示] 商品图片目录不存在：{PRODUCT_IMAGE_DIR}，销售排行将显示缺图占位")

# =========================================================
# 3. 库存数据
# =========================================================
print("[2/6] 读取库存数据...")
WH_TO_CHAN = {
    "达芙妮DL电商仓":"DL线上","达芙妮DL广东仓":"DL线上","达芙妮DL微瑕仓":"DL线上",
    "达芙妮DL门店总仓":"DL线下",
    "DAPHNE.LAB成都COSMO直营店":"DL线下","DAPHNE.LAB南京艾尚天地直营店":"DL线下",
    "DAPHNE.LAB上海TX淮海直营店":"DL线下","DAPHNE.LAB上海淮海百盛直营店":"DL线下",
    "DAPHNE.LAB长沙TX直营店":"DL线下",
    "DAPHNE上海奉贤金汇天街直营店":"DA","DAPHNE上海普陀中环百联直营店":"DA",
    "达芙妮DA直营仓":"DA",
}

# 以映射表A/B列为最新仓库渠道映射。
mapping_wb = openpyxl.load_workbook(path("仓库与店铺对应关系表.xlsx"), read_only=True, data_only=True)
mapping_ws = mapping_wb["Sheet1"]
for r in mapping_ws.iter_rows(min_row=2, values_only=True):
    warehouse = r[0] if len(r) > 0 else None
    channel = r[1] if len(r) > 1 else None
    if warehouse and channel:
        WH_TO_CHAN[str(warehouse).strip()] = str(channel).strip()
mapping_wb.close()

wb = openpyxl.load_workbook(path("库存数据.xlsx"), read_only=True, data_only=True)
ws = wb["Sheet1"]

# 5大类库存
inv_cat   = defaultdict(int)   # cat -> qty
inv_nianji = defaultdict(int)  # nianji -> qty
inv_small  = defaultdict(int)  # small -> qty
inv_small_drill = defaultdict(lambda: defaultdict(int))  # small -> nianji -> qty
inv_wh     = defaultdict(int)  # wh -> qty (5大类)
inv_wh_cat = defaultdict(lambda: defaultdict(int))  # cat -> wh -> qty
inv_cat_nianji = defaultdict(lambda: defaultdict(int))  # cat -> nianji -> qty
inv_channel_agg = {
    ch: {
        "cat": defaultdict(int), "nianji": defaultdict(int), "small": defaultdict(int),
        "smallDrill": defaultdict(lambda: defaultdict(int)), "wh": defaultdict(lambda: defaultdict(int)),
        "catNianji": defaultdict(lambda: defaultdict(int)),
        "catSmall": defaultdict(lambda: defaultdict(int)),
        "catSmallDrill": defaultdict(lambda: defaultdict(lambda: defaultdict(int))),
    }
    for ch in ("all", "DL线上", "DL线下")
}
# SKC×渠道 库存（3大类）
skc_inv_chan = defaultdict(lambda: defaultdict(int))  # skc -> chan -> qty
excluded_payu_rows = 0
excluded_payu_qty = 0

for r in ws.iter_rows(min_row=2, values_only=True):
    skc, size, wh, qty = r
    qty = qty or 0
    # 库存口径统一排除所有“帕羽”仓库（如南昌帕羽）。
    if wh and "帕羽" in str(wh):
        excluded_payu_rows += 1
        excluded_payu_qty += qty
        continue
    meta = skc_meta.get(skc)
    if not meta: continue
    cat = meta["cat"]
    inv_cat[cat] += qty
    inv_nianji[meta["nianji"]] += qty
    inv_small[meta["small"]] += qty
    inv_small_drill[meta["small"]][meta["nianji"]] += qty
    inv_wh[wh] += qty
    inv_wh_cat[cat][wh] += qty
    inv_cat_nianji[cat][meta["nianji"]] += qty
    chan = WH_TO_CHAN.get(wh, "all")
    for view_chan in (["all", chan] if chan in {"DL线上", "DL线下"} else ["all"]):
        agg = inv_channel_agg[view_chan]
        agg["cat"][cat] += qty
        agg["nianji"][meta["nianji"]] += qty
        agg["small"][meta["small"]] += qty
        agg["smallDrill"][meta["small"]][meta["nianji"]] += qty
        agg["wh"][cat][wh] += qty
        agg["catNianji"][cat][meta["nianji"]] += qty
        agg["catSmall"][cat][meta["small"]] += qty
        agg["catSmallDrill"][cat][meta["small"]][meta["nianji"]] += qty
    if cat in TARGET_CATS_5:
        skc_inv_chan[skc][chan] += qty

# 序列化
CAT_ORDER = ["女鞋","服装","箱包","配件","赠品"]
inv_cat_json = {c: inv_cat.get(c, 0) for c in CAT_ORDER}
inv_cat_nianji_json = {c: dict(inv_cat_nianji[c]) for c in CAT_ORDER}

def serialize_inv_channel(agg):
    small_sorted = sorted(agg["small"].items(), key=lambda x: -x[1])
    small_top = [{"label": k, "v": v} for k, v in small_sorted[:15]]
    small_rest = sum(v for _, v in small_sorted[15:])
    if small_rest > 0: small_top.append({"label": "其他小类", "v": small_rest})
    cat_small = {}
    for cat in CAT_ORDER:
        items = sorted(agg["catSmall"][cat].items(), key=lambda x: -x[1])
        top = [{"label": k, "v": v} for k, v in items[:15]]
        rest = sum(v for _, v in items[15:])
        if rest > 0: top.append({"label": "其他小类", "v": rest})
        cat_small[cat] = top
    return {
        "cat": {c: agg["cat"].get(c, 0) for c in CAT_ORDER},
        "nianji": [{"label": k, "v": v} for k, v in agg["nianji"].items()],
        "small": small_top,
        "smallDrill": {sm: [{"label": k, "v": v} for k, v in nd.items()] for sm, nd in agg["smallDrill"].items()},
        "wh": {c: {w: q for w, q in sorted(agg["wh"][c].items(), key=lambda x: -x[1])[:10]} for c in CAT_ORDER},
        "catNianji": {c: dict(agg["catNianji"][c]) for c in CAT_ORDER},
        "catSmall": cat_small,
        "catSmallDrill": {
            c: {sm: [{"label": k, "v": v} for k, v in nd.items()] for sm, nd in agg["catSmallDrill"][c].items()}
            for c in CAT_ORDER
        },
    }

inv_channel_data = {ch: serialize_inv_channel(agg) for ch, agg in inv_channel_agg.items()}

nianji_sorted = sorted(inv_nianji.items(), key=lambda x: (x[0][:2], SEASON_ORDER.get(x[0][2:], 99)))
inv_nianji_json = [{"label": k, "v": v} for k, v in nianji_sorted]

small_sorted = sorted(inv_small.items(), key=lambda x: -x[1])
TOP_N = 15
inv_small_json = [{"label": k, "v": v} for k, v in small_sorted[:TOP_N]]
rest_sum = sum(v for _, v in small_sorted[TOP_N:])
if rest_sum > 0: inv_small_json.append({"label": "其他小类", "v": rest_sum})

inv_small_drill_json = {
    sm: [{"label": k, "v": v} for k, v in sorted(nd.items(), key=lambda x: (x[0][:2], SEASON_ORDER.get(x[0][2:], 99)))]
    for sm, nd in inv_small_drill.items()
}

wh_sorted = sorted(inv_wh.items(), key=lambda x: -x[1])[:10]

skc_inv_chan_json = {skc: dict(d) for skc, d in skc_inv_chan.items()}

# 总库存量（用于模板中显示）
total_5cat_inv = sum(inv_cat.values())
print(f"    5大类库存合计：{total_5cat_inv:,} 件")
print(f"    已排除帕羽仓库：{excluded_payu_rows:,} 行，{excluded_payu_qty:,} 件")

# =========================================================
# 4. 销售数据
# =========================================================
print("[3/6] 读取销售数据（较大，请稍候）...")
wb = openpyxl.load_workbook(path("销售数据.xlsx"), read_only=True, data_only=True)
ws = wb["Sheet1"]

# 找到数据截止日期（用于近28天计算）
all_dates = []
daily_agg = defaultdict(lambda: [0]*13)
sku_day_chan = defaultdict(lambda: [0]*5)
row_count = 0

for r in ws.iter_rows(min_row=2, values_only=True):
    row_count += 1
    store = r[5]; cust = STORE_MAPPING.get(store, "其他")
    date = r[1]; date = date if isinstance(date, str) else str(date)
    date = date.replace("/","-")
    parts = date.split("-")
    if len(parts) == 3:
        y, m, d = parts; date = f"{y}-{int(m):02d}-{int(d):02d}"
    all_dates.append(date)

    # daily_agg (收入/销售概况)
    key = (date, cust); a = daily_agg[key]
    net_qty = r[6] or 0; net_amt = r[7] or 0
    sale_qty = r[8] or 0; sale_amt = r[9] or 0
    a[0]+=net_qty; a[1]+=net_amt; a[2]+=sale_qty; a[3]+=sale_amt
    a[4]+=r[10] or 0; a[5]+=r[11] or 0
    a[6]+=r[15] or 0; a[7]+=r[16] or 0; a[8]+=r[17] or 0; a[9]+=r[18] or 0
    if net_amt == 0 and net_qty > 0: a[10] += net_qty
    if sale_amt == 0 and sale_qty > 0: a[11] += sale_qty
    # 件单价分母需排除实发金额为0的商品实发数量。
    ship_qty = r[15] or 0; ship_amt = r[16] or 0
    if ship_amt == 0 and ship_qty > 0: a[12] += ship_qty

    # sku_day_chan (销售排行 + 工厂图表)
    skc = r[3]
    if skc in skc_meta and skc_meta[skc]["cat"] in TARGET_CATS_5:
        chan = chan_of(cust)
        key2 = (date, skc, chan); b = sku_day_chan[key2]
        b[0]+=r[6] or 0; b[1]+=r[8] or 0; b[2]+=r[10] or 0
        b[3]+=r[15] or 0; b[4]+=r[17] or 0

max_date = max(all_dates) if all_dates else "2026-06-29"
min_date = min(all_dates) if all_dates else "2025-01-01"
print(f"    销售数据 {row_count:,} 行，日期范围：{min_date} ~ {max_date}")

# 序列化 daily_agg
daily_out = sorted([[date,cust]+[round(x,2) for x in v] for (date,cust),v in daily_agg.items()])

# sku_month_chan_agg
month_agg = defaultdict(lambda: [0]*5)
for (date, skc, chan), v in sku_day_chan.items():
    ym = date[:7]; key = (ym, skc, chan); a = month_agg[key]
    for i in range(5): a[i] += v[i]
sku_month_chan = sorted([[ym,skc,chan]+v for (ym,skc,chan),v in month_agg.items()])
sku_day_chan_out = sorted([[date,skc,chan]+v for (date,skc,chan),v in sku_day_chan.items()])

# 近28天（截至 max_date 的前28天）
from datetime import datetime, timedelta
max_dt = datetime.strptime(max_date, "%Y-%m-%d")
last28_start = (max_dt - timedelta(days=27)).strftime("%Y-%m-%d")
sku_last28 = [
    [date,skc,chan]+v for (date,skc,chan),v in sku_day_chan.items()
    if date >= last28_start and date <= max_date
]
print(f"    近4周区间：{last28_start} ~ {max_date}，记录数：{len(sku_last28)}")

# =========================================================
# 5. 生产工厂数量占比数据（SKC级别，供JS动态计算）
# =========================================================
# 已通过 skc_factory, skc_inv_chan, SKU_MONTH_CHAN 在JS里动态计算，这里不需要预聚合

# =========================================================
# 6. 库存概况模块更新仓库总览HTML（动态写入最新的仓库排名）
# =========================================================
wh_bars_by_cat = {
    cat: {wh: qty for wh, qty in sorted(inv_wh_cat[cat].items(), key=lambda x: -x[1])[:10]}
    for cat in CAT_ORDER
}

# =========================================================
# 7. 指标达成目标（来自指标.xlsx）
# =========================================================
print("[4/6] 读取指标数据...")
wb = openpyxl.load_workbook(path("指标.xlsx"), read_only=True, data_only=True)
ws = wb["Sheet1"]
targets = {}
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r[0]: continue
    cust = str(r[0]).strip()
    monthly = [float(v) if v else 0 for v in r[1:13]]
    if len(monthly) == 12 and any(monthly):
        targets[cust] = monthly

print(f"    指标客户数：{len(targets)}，客户：{list(targets.keys())}")

# =========================================================
# 8. 组装数据
# =========================================================
print("[5/6] 组装数据...")

def top_n_other(items_dict, n=8):
    sorted_items = sorted(items_dict.items(), key=lambda x: -x[1])
    top = sorted_items[:n]
    rest = sum(v for _, v in sorted_items[n:])
    result = [{"label": k, "v": v} for k, v in top]
    if rest > 0: result.append({"label": "其他工厂", "v": rest})
    return result

# =========================================================
# 9. 注入模板
# =========================================================
print("[6/6] 注入模板，生成看板...")

with open(path(TEMPLATE), encoding="utf-8") as f:
    content = f.read()

def inject(content, placeholder, data):
    assert placeholder in content, f"模板中找不到占位符：{placeholder}"
    js = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
    return content.replace(placeholder, js)

# 替换动态最新日期文字
content = content.replace("__MAX_DATE__", max_date)
content = content.replace("__MIN_DATE__", min_date)
content = content.replace("__MAX_MONTH__", max_date[:7])
content = content.replace("__MIN_MONTH__", min_date[:7])
content = content.replace("__CURRENT_MONTH_START__", max_date[:7] + "-01")
content = content.replace("__LAST28_START__", last28_start)
content = content.replace("__MAX_YEAR__", max_date[:4])

content = inject(content, "__DAILY_DATA__",       daily_out)
content = inject(content, "__TARGETS__",           targets)
content = inject(content, "__SKU_MONTH_CHAN__",     sku_month_chan)
content = inject(content, "__SKU_DAY_CHAN__",       sku_day_chan_out)
content = inject(content, "__SKU_LAST28_CHAN__",    sku_last28)
content = inject(content, "__SKC_CAT__",           skc_cat)
content = inject(content, "__SKC_NAME__",          skc_name)
content = inject(content, "__SKC_IMAGE__",         skc_image)
content = inject(content, "__SKC_INV_CHAN__",      skc_inv_chan_json)
content = inject(content, "__SKC_FACTORY__",       skc_factory)
content = inject(content, "__SKC_NIANJI__",       skc_nianji)
content = inject(content, "__INV_CAT_NIANJI__",   inv_cat_nianji_json)
content = inject(content, "__INV_CHANNEL_DATA__", inv_channel_data)

output_path = path("index.html")
with open(output_path, "w", encoding="utf-8") as f:
    f.write(content)

print("=" * 60)
print(f"[OK] 生成完成：index.html")
print(f"   数据范围：{min_date} ~ {max_date}")
print(f"   文件大小：{os.path.getsize(output_path)/1024:.0f} KB")
print("=" * 60)
